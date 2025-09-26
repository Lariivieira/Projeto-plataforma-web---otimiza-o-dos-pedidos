# app.py
import os
from flask import Flask, request, render_template, send_file, redirect, url_for, flash
from werkzeug.utils import secure_filename
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

UPLOAD_FOLDER = 'uploads'
OUTPUT_FOLDER = 'outputs'
ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv'}

for d in (UPLOAD_FOLDER, OUTPUT_FOLDER):
    os.makedirs(d, exist_ok=True)

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.secret_key = 'troque_isto_para_algo_secreto'

# -------------------------
# Utilitários
# -------------------------
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def load_multiples(path='multiplos.csv'):
    import pandas as pd
    try:
        df = pd.read_csv(path, sep=';', dtype=str).fillna('')  # tenta ponto e vírgula
    except pd.errors.ParserError:
        df = pd.read_csv(path, sep=',', dtype=str).fillna('')  # tenta vírgula

    # remove espaços extras dos nomes das colunas
    df.columns = [c.strip() for c in df.columns]

    # força os nomes esperados
    if "Número do modelo" not in df.columns or "Quantidade solicitada" not in df.columns:
        raise ValueError("A planilha de múltiplos precisa ter as colunas 'Número do modelo' e 'Quantidade solicitada'.")

    d = {}
    for _, row in df.iterrows():
        try:
            sku = str(row["Número do modelo"]).strip()
            mult = int(float(str(row["Quantidade solicitada"]).strip()))
            d[sku] = mult
        except:
            continue

    return d

def detect_columns(df):
    cols = {c.lower(): c for c in df.columns}
    sku_candidates = ['sku','número do modelo','codigo','cod','ean','gtin','produto','product']
    qty_candidates = ['quantity','qty','qtd','quantidade','quantidade solicitada','amount','amountordered']
    sku_col = next((cols[c] for c in sku_candidates if c in cols), None)
    qty_col = next((cols[c] for c in qty_candidates if c in cols), None)
    return sku_col, qty_col

def process_dataframe(df, multiples):
    sku_col, qty_col = detect_columns(df)
    if sku_col is None or qty_col is None:
        raise ValueError("Não foi possível detectar colunas SKU/Quantidade automaticamente. Verifique o arquivo.")
    df = df.copy()
    df['_SKU_'] = df[sku_col].astype(str).str.strip()
    df['_QTY_'] = pd.to_numeric(df[qty_col], errors='coerce').fillna(0).astype(int)

    statuses, adj_qtys = [], []
    for _, row in df.iterrows():
        sku = row['_SKU_']
        qty = int(row['_QTY_'])
        mult = multiples.get(sku)
        if mult is None:
            statuses.append('Este_item_não_existe_na_planilha_de_múltiplo_atualize')
            adj_qtys.append(qty)
        else:
            if qty < mult:
                statuses.append('abaixo_do_mínimo')
                adj_qtys.append(0)
            else:
                adj = (qty // mult) * mult
                if adj != qty:
                    statuses.append('ajustado')
                    adj_qtys.append(adj)
                else:
                    statuses.append('ok')
                    adj_qtys.append(qty)

    df['Status'] = statuses
    df['AdjustedQty'] = adj_qtys

    pedido_df = df[df['Status'] != 'abaixo_do_mínimo'].copy()
    pedido_df[qty_col] = pedido_df['AdjustedQty']
    controle_df = df[df['Status'] == 'abaixo_do_mínimo'].copy()
    missing_df = df[df['Status'] == 'Este_item_não_existe_na_planilha_de_múltiplo_atualize'].copy()
    controle_df = pd.concat([controle_df, missing_df], ignore_index=True)

    # Excluir colunas indesejadas
    unwanted_cols = ["Custo total", "Unnamed: 15", "Quantidade aceita", "Quantidade de ASN", "Quantidade recebida", "Quantidade pendente", "SKU", "AdjustedQty", "_QTY_", "_SKU_"]
    for df_out in [pedido_df, controle_df]:
        df_out.drop(columns=[col for col in unwanted_cols if col in df_out.columns], inplace=True, errors='ignore')

    # Renomear colunas
    for df_out in [pedido_df, controle_df]:
        if 'QTY' in df_out.columns:
            df_out.rename(columns={'QTY': 'Quantidade solicitada pela Amazon'}, inplace=True)
        if 'Quantidade solicitada' in df_out.columns:
            df_out.rename(columns={'Quantidade solicitada': 'Quantidade Ajustada'}, inplace=True)

    return pedido_df, controle_df, sku_col, qty_col

def style_excel(path, status_col_name='Status'):
    wb = load_workbook(path)
    sheets = ['Pedido', 'Removidos']
    fill_adjusted = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")  # amarelo
    fill_below = PatternFill(start_color="FFBABA", end_color="FFBABA", fill_type="solid")     # vermelho claro
    fill_missing = PatternFill(start_color="FFD8A8", end_color="FFD8A8", fill_type="solid")   # laranja

    for sheet_name in sheets:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            if status_col_name not in headers:
                continue
            status_idx = headers.index(status_col_name) + 1

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                status = row[status_idx - 1].value
                if status == 'ajustado':
                    for cell in row:
                        cell.fill = fill_adjusted
                elif status == 'abaixo_do_mínimo':
                    for cell in row:
                        cell.fill = fill_below
                elif status == 'Este_item_não_existe_na_planilha_de_múltiplo_atualize':
                    for cell in row:
                        cell.fill = fill_missing

    wb.save(path)

# -------------------------
# Rotas
# -------------------------
@app.route('/', methods=['GET','POST'])
def index():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('Nenhum arquivo enviado')
            return redirect(request.url)
        f = request.files['file']
        if f.filename == '':
            flash('Nenhum arquivo selecionado')
            return redirect(request.url)
        if f and allowed_file(f.filename):
            filename = secure_filename(f.filename)
            in_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            f.save(in_path)

            multiples = load_multiples('multiplos.csv')

            try:
                if filename.lower().endswith('.csv'):
                    df = pd.read_csv(in_path)
                else:
                    df = pd.read_excel(in_path, engine='openpyxl')
            except Exception as e:
                flash(f'Erro ao ler o arquivo: {e}')
                return redirect(request.url)

            try:
                pedido_df, controle_df, sku_col, qty_col = process_dataframe(df, multiples)
            except Exception as e:
                flash(str(e))
                return redirect(request.url)

            order_name = request.form.get('order_name') or os.path.splitext(filename)[0]
            out_filename = f"{secure_filename(order_name)}_processado.xlsx"
            out_path = os.path.join(OUTPUT_FOLDER, out_filename)
            with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
                pedido_df.to_excel(writer, sheet_name='Pedido', index=False)
                controle_df.to_excel(writer, sheet_name='Removidos', index=False)

            style_excel(out_path)

            return render_template(
                'resultado.html',
                download_url=url_for('download_file', filename=out_filename),
                filename=out_filename
            )

    return render_template('index.html')

@app.route('/download/<filename>')
def download_file(filename):
    path = os.path.join(OUTPUT_FOLDER, filename)
    return send_file(path, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
